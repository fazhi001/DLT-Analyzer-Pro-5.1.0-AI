import numpy as np

from dlt_analyzer_pro.credible_evaluation import (
    brier_skill_score,
    evaluate_zone_credibility,
)


def test_brier_skill_score_direction():
    assert brier_skill_score(0.10, 0.20) == 0.5
    assert brier_skill_score(0.20, 0.20) == 0.0
    assert brier_skill_score(0.30, 0.20) < 0.0


def test_bootstrap_and_random_baseline_are_reproducible():
    model_briers = np.array([0.10, 0.11, 0.09, 0.12, 0.10, 0.11])
    reference = np.full(6, 0.122449)
    hits = np.array([1, 1, 2, 0, 1, 1])
    first = evaluate_zone_credibility(
        model_briers, reference, hits, pool_size=35, pick_count=5,
        bootstrap_samples=500, random_repeats=1000, seed=42,
    )
    second = evaluate_zone_credibility(
        model_briers, reference, hits, pool_size=35, pick_count=5,
        bootstrap_samples=500, random_repeats=1000, seed=42,
    )
    assert first == second
    assert first.brier_skill_score > 0
    assert first.bss_ci_lower <= first.brier_skill_score <= first.bss_ci_upper
    assert 0.0 <= first.random_p_value <= 1.0


def test_random_baseline_mean_matches_theory():
    result = evaluate_zone_credibility(
        [0.122] * 20, [0.122449] * 20, [1] * 20,
        pool_size=35, pick_count=5, bootstrap_samples=500,
        random_repeats=10000, seed=7,
    )
    expected = 25 / 35
    assert abs(result.random_hit_average - expected) < 0.03


def test_credible_evaluation_excel_export(tmp_path):
    from openpyxl import load_workbook
    from dlt_analyzer_pro.ai_backtest import walk_forward_ai_backtest
    from dlt_analyzer_pro.exporter import export_backtest_evaluation_xlsx
    from dlt_analyzer_pro.importer import load_file
    from dlt_analyzer_pro.paths import resource_path

    draws, failures = load_file(resource_path("dlt_history.csv"))
    assert not failures
    result = walk_forward_ai_backtest(
        draws,
        periods=3,
        include_ml=False,
        bootstrap_samples=500,
        random_repeats=1000,
    )
    target = tmp_path / "credible.xlsx"
    export_backtest_evaluation_xlsx(target, result)
    workbook = load_workbook(target, read_only=True)
    assert workbook.sheetnames == ["可信评估摘要", "逐期回测", "方法说明"]
