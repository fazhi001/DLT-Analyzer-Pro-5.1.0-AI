from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook

from .models import DigitBacktestResult, DigitDraw, DigitPrediction


GAME_NAMES = {"pl3": "排列三", "pl5": "排列五"}


def export_digit_draws_xlsx(path: Path, draws: Iterable[DigitDraw]) -> None:
    draws = list(draws)
    game = draws[0].game if draws else "pl5"
    count = 3 if game == "pl3" else 5
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = GAME_NAMES.get(game, game)
    sheet.append(["期号", "开奖日期", *[f"第{i}位" for i in range(1, count + 1)], "开奖号码"])
    for draw in draws:
        sheet.append([
            draw.issue,
            draw.draw_date.isoformat() if draw.draw_date else "",
            *draw.digits,
            draw.number_text if hasattr(draw, "number_text") else "".join(map(str, draw.digits)),
        ])
    sheet.freeze_panes = "A2"
    workbook.save(path)


def export_digit_predictions_xlsx(
    path: Path,
    game: str,
    target_issue: str,
    predictions: Iterable[DigitPrediction],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "预测结果"
    sheet.append(["玩法", "目标期号", "序号", "号码", "评分", "策略", "模型模式", "和值", "跨度"])
    for index, item in enumerate(predictions, start=1):
        sheet.append([
            GAME_NAMES.get(game, game), target_issue, index, item.number_text,
            item.score, item.strategy, item.model_mode,
            sum(item.digits), max(item.digits) - min(item.digits),
        ])
    workbook.save(path)


def export_digit_backtest_xlsx(path: Path, result: DigitBacktestResult) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "回测汇总"
    summary.append(["玩法", GAME_NAMES.get(result.game, result.game)])
    summary.append(["验证期数", result.evaluated])
    summary.append(["模型平均命中位置", result.model_average_hits])
    summary.append(["随机平均命中位置", result.random_average_hits])
    summary.append(["模型完整命中", result.model_exact_hits])
    summary.append(["随机完整命中", result.random_exact_hits])
    summary.append([])
    summary.append(["位置", "模型命中率", "随机命中率"])
    for index, (model, random_rate) in enumerate(
        zip(result.position_model_rates, result.position_random_rates), start=1
    ):
        summary.append([index, model, random_rate])

    detail = workbook.create_sheet("逐期明细")
    detail.append(["期号", "模型命中位置", "随机命中位置", "模型完整命中", "随机完整命中"])
    for item in result.details:
        detail.append([
            item.issue, item.model_hits, item.random_hits,
            "是" if item.exact_model else "否", "是" if item.exact_random else "否",
        ])
    workbook.save(path)
