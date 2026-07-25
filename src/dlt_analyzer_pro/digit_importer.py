from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .models import DigitDraw


ISSUE_ALIASES = {"issue", "期号", "期次", "lotterydrawnum"}
DATE_ALIASES = {"date", "draw_date", "开奖日期", "日期", "lotterydrawtime"}
NUMBER_ALIASES = {"number", "numbers", "开奖号码", "开奖号", "lotterydrawresult"}
PL3_POSITION_ALIASES = [
    ("d1", "百位", "第一位", "位置1"),
    ("d2", "十位", "第二位", "位置2"),
    ("d3", "个位", "第三位", "位置3"),
]
PL5_POSITION_ALIASES = [
    ("d1", "万位", "第一位", "位置1"),
    ("d2", "千位", "第二位", "位置2"),
    ("d3", "百位", "第三位", "位置3"),
    ("d4", "十位", "第四位", "位置4"),
    ("d5", "个位", "第五位", "位置5"),
]


def _clean(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "")


def _find(headers: list[str], aliases) -> str | None:
    normalized = {_clean(header): header for header in headers}
    for alias in aliases:
        if _clean(alias) in normalized:
            return normalized[_clean(alias)]
    return None


def _parse_date(value: object):
    if value in (None, ""):
        return None
    if hasattr(value, "date") and not isinstance(value, str):
        try:
            return value.date()
        except Exception:
            pass
    text = str(value).strip()[:10]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"无法识别开奖日期：{value}")


def _parse_digits(value: object, expected: int) -> list[int]:
    if value is None:
        return []
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        text = str(int(value)).zfill(expected)
    else:
        text = str(value).strip()
    separated = re.findall(r"\d+", text)
    if len(separated) in (3, 5) and all(len(item) == 1 for item in separated):
        return [int(item) for item in separated]
    compact = "".join(separated)
    if len(compact) in (3, 5):
        return [int(char) for char in compact]
    return []


def _row_to_draw(game: str, row: dict[str, object], headers: list[str]) -> DigitDraw:
    issue_column = _find(headers, ISSUE_ALIASES)
    date_column = _find(headers, DATE_ALIASES)
    number_column = _find(headers, NUMBER_ALIASES)
    if issue_column is None:
        raise ValueError("缺少期号列")
    raw_issue = row.get(issue_column, "")
    if isinstance(raw_issue, (int, float)) and not isinstance(raw_issue, bool):
        issue = str(int(raw_issue))
    else:
        issue = str(raw_issue).strip()
    expected = 3 if game == "pl3" else 5

    digits: list[int] = []
    if number_column:
        digits = _parse_digits(row.get(number_column), expected)
    if not digits:
        position_aliases = PL3_POSITION_ALIASES if game == "pl3" else PL5_POSITION_ALIASES
        for aliases in position_aliases:
            column = _find(headers, aliases)
            if column is None:
                digits = []
                break
            digits.append(int(row.get(column)))
    if game == "pl3" and len(digits) == 5:
        digits = digits[:3]
    if len(digits) != expected:
        raise ValueError(f"开奖号码应为{expected}位")

    draw = DigitDraw(
        game=game,
        issue=issue,
        draw_date=_parse_date(row.get(date_column)) if date_column else None,
        digits=tuple(digits),
    )
    draw.validate()
    return draw


def load_digit_file(path: Path, game: str) -> tuple[list[DigitDraw], list[dict[str, object]]]:
    path = Path(path)
    game = game.lower()
    if game not in {"pl3", "pl5"}:
        raise ValueError(f"不支持的玩法：{game}")

    rows: list[dict[str, object]] = []
    if path.suffix.lower() == ".csv":
        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    rows = list(csv.DictReader(handle))
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        else:
            raise ValueError("CSV字符编码无法识别") from last_error
    elif path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return [], []
        headers = [str(value or "").strip() for value in values[0]]
        rows = [dict(zip(headers, row)) for row in values[1:]]
    else:
        raise ValueError("仅支持 CSV、XLSX、XLSM 文件")

    if not rows:
        return [], []
    headers = list(rows[0].keys())
    draws: list[DigitDraw] = []
    failures: list[dict[str, object]] = []
    for index, row in enumerate(rows, start=2):
        try:
            draws.append(_row_to_draw(game, row, headers))
        except Exception as exc:
            failures.append({"row": index, "error": str(exc), "data": row})
    return draws, failures
