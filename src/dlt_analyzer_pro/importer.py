from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import Draw


ALIASES = {
    "issue": ("issue", "期号", "draw_num", "draw_number"),
    "date": ("date", "开奖日期", "draw_date"),
    "f1": ("f1", "前区1", "red1"),
    "f2": ("f2", "前区2", "red2"),
    "f3": ("f3", "前区3", "red3"),
    "f4": ("f4", "前区4", "red4"),
    "f5": ("f5", "前区5", "red5"),
    "b1": ("b1", "后区1", "blue1"),
    "b2": ("b2", "后区2", "blue2"),
}


def _resolve_headers(headers: Iterable[str]) -> dict[str, str]:
    normalized = {str(h).strip().lower(): str(h) for h in headers if h is not None}
    result: dict[str, str] = {}
    for field, candidates in ALIASES.items():
        for candidate in candidates:
            if candidate.lower() in normalized:
                result[field] = normalized[candidate.lower()]
                break
    required = {"issue", "f1", "f2", "f3", "f4", "f5", "b1", "b2"}
    missing = required - result.keys()
    if missing:
        raise ValueError("数据文件缺少字段：" + "、".join(sorted(missing)))
    return result


def _parse_date(value: object) -> datetime.date | None:
    if value is None:
        return None
    if hasattr(value, "date") and not isinstance(value, str):
        try:
            return value.date()
        except Exception:
            pass
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"无法识别日期：{text}")


def _row_to_draw(row: dict[str, object], columns: dict[str, str]) -> Draw:
    front = tuple(
        sorted(int(row[columns[f"f{i}"]]) for i in range(1, 6))
    )
    back = tuple(
        sorted(int(row[columns[f"b{i}"]]) for i in range(1, 3))
    )
    draw = Draw(
        issue=str(row[columns["issue"]]).strip(),
        draw_date=_parse_date(row.get(columns.get("date", ""))),
        front=front,
        back=back,
    )
    draw.validate()
    return draw


def load_csv(path: Path) -> tuple[list[Draw], list[dict[str, object]]]:
    draws: list[Draw] = []
    failures: list[dict[str, object]] = []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            raise ValueError("CSV为空或没有表头")
        columns = _resolve_headers(reader.fieldnames)
        for line, row in enumerate(reader, start=2):
            try:
                draws.append(_row_to_draw(dict(row), columns))
            except Exception as exc:
                failures.append({"line": line, "error": str(exc)})
    return draws, failures


def load_xlsx(path: Path) -> tuple[list[Draw], list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(v).strip() if v is not None else "" for v in next(rows)]
    except StopIteration as exc:
        raise ValueError("XLSX为空") from exc
    columns = _resolve_headers(headers)
    draws: list[Draw] = []
    failures: list[dict[str, object]] = []
    for line, values in enumerate(rows, start=2):
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        try:
            draws.append(_row_to_draw(row, columns))
        except Exception as exc:
            failures.append({"line": line, "error": str(exc)})
    workbook.close()
    return draws, failures


def load_file(path: Path) -> tuple[list[Draw], list[dict[str, object]]]:
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx(path)
    raise ValueError("仅支持 CSV 或 XLSX 文件")
