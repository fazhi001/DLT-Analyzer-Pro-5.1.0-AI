from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .analytics import draw_metrics
from .models import Draw, Prediction


def export_draws_csv(path: Path, draws: Iterable[Draw]) -> None:
    with Path(path).open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["期号", "开奖日期", "前区1", "前区2", "前区3", "前区4", "前区5", "后区1", "后区2"])
        for draw in draws:
            writer.writerow(
                [
                    draw.issue,
                    draw.draw_date.isoformat() if draw.draw_date else "",
                    *draw.front,
                    *draw.back,
                ]
            )


def export_predictions_xlsx(
    path: Path,
    target_issue: str,
    predictions: Iterable[Prediction],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "预测结果"
    sheet.append(["目标期号", "序号", "策略", "前区", "后区", "相对综合分"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for index, prediction in enumerate(predictions, start=1):
        sheet.append(
            [
                target_issue,
                index,
                prediction.strategy,
                " ".join(f"{n:02d}" for n in prediction.front),
                " ".join(f"{n:02d}" for n in prediction.back),
                prediction.score,
            ]
        )
    widths = {"A": 14, "B": 8, "C": 14, "D": 28, "E": 14, "F": 12}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    workbook.save(path)


def export_draws_xlsx(path: Path, draws: Iterable[Draw]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "历史开奖"
    sheet.append(["期号", "开奖日期", "前区", "后区", "和值", "奇偶", "分区"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for draw in draws:
        metrics = draw_metrics(draw)
        sheet.append(
            [
                draw.issue,
                draw.draw_date.isoformat() if draw.draw_date else "",
                " ".join(f"{n:02d}" for n in draw.front),
                " ".join(f"{n:02d}" for n in draw.back),
                metrics["sum"],
                f'{metrics["odd"]}:{metrics["even"]}',
                ":".join(str(n) for n in metrics["zones"]),
            ]
        )
    for column, width in {
        "A": 12,
        "B": 14,
        "C": 28,
        "D": 14,
        "E": 10,
        "F": 10,
        "G": 10,
    }.items():
        sheet.column_dimensions[column].width = width
    workbook.save(path)


def export_ai_report_xlsx(
    path: Path,
    target_issue: str,
    predictions: Iterable[Prediction],
    report,
) -> None:
    workbook = Workbook()

    result_sheet = workbook.active
    result_sheet.title = "AI预测结果"
    result_sheet.append(
        ["目标期号", "序号", "前区", "后区", "相对综合分", "模型"]
    )
    for cell in result_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for index, prediction in enumerate(predictions, start=1):
        result_sheet.append(
            [
                target_issue,
                index,
                " ".join(f"{n:02d}" for n in prediction.front),
                " ".join(f"{n:02d}" for n in prediction.back),
                prediction.score,
                prediction.strategy,
            ]
        )

    model_sheet = workbook.create_sheet("模型验证")
    model_sheet.append(["模型", "区域", "训练样本", "校准前Brier", "校准后Brier", "验证AUC", "校准方法", "缓存状态", "模型版本"])
    for cell in model_sheet[1]:
        cell.font = Font(bold=True)
    for metric in report.model_metrics:
        model_sheet.append(
            [
                metric.model_name,
                metric.zone,
                metric.train_rows,
                getattr(metric, "uncalibrated_brier", None),
                metric.validation_brier,
                metric.validation_auc,
                getattr(metric, "calibration_method", "identity"),
                getattr(metric, "cache_status", "trained"),
                getattr(metric, "model_version", ""),
            ]
        )

    weight_sheet = workbook.create_sheet("集成权重")
    weight_sheet.append(["模型组件", "综合权重", "前区权重", "后区权重", "来源"])
    for cell in weight_sheet[1]:
        cell.font = Font(bold=True)
    front_weights = getattr(report, "front_component_weights", {}) or report.component_weights
    back_weights = getattr(report, "back_component_weights", {}) or report.component_weights
    for name, value in report.component_weights.items():
        weight_sheet.append([
            name, value, front_weights.get(name), back_weights.get(name),
            getattr(report, "weight_source", "manual"),
        ])
    weight_sheet.append(["均匀概率基线", None, 1.0 - getattr(report, "front_model_share", 1.0), 1.0 - getattr(report, "back_model_share", 1.0), "基线保护"])

    dynamic = getattr(report, "dynamic_result", None)
    if dynamic is not None:
        rank_sheet = workbook.create_sheet("动态模型排行榜")
        rank_sheet.append(["排名", "模型", "最终权重", "权重变化", "前区命中", "后区命中", "前区Brier", "后区Brier", "质量分"])
        for rank, item in enumerate(dynamic.rankings, start=1):
            rank_sheet.append([rank, item.label, item.final_weight, item.trend, item.front_hits, item.back_hits, item.front_brier, item.back_brier, item.quality_score])

    front_sheet = workbook.create_sheet("前区概率评分")
    front_sheet.append(["号码", "集成概率"])
    for number, value in report.front_scores.items():
        front_sheet.append([number, value])

    back_sheet = workbook.create_sheet("后区概率评分")
    back_sheet.append(["号码", "集成概率"])
    for number, value in report.back_scores.items():
        back_sheet.append([number, value])

    info_sheet = workbook.create_sheet("运行信息")
    info_sheet.append(["最近五年样本期数", report.dataset_count])
    info_sheet.append(["蒙特卡洛模拟次数", report.simulations])
    info_sheet.append(["运行耗时（秒）", report.elapsed_seconds])
    info_sheet.append(["固定随机种子", getattr(report, "deterministic_seed", 0)])
    info_sheet.append(["数据指纹", getattr(report, "dataset_fingerprint", "")])
    info_sheet.append(["数据泄漏检查", "通过" if getattr(report, "leakage_audit_passed", True) else "未通过"])
    info_sheet.append(["前区模型概率占比", getattr(report, "front_model_share", 1.0)])
    info_sheet.append(["后区模型概率占比", getattr(report, "back_model_share", 1.0)])
    for note in getattr(report, "baseline_guard_notes", ()):
        info_sheet.append(["基线保护", note])
    for note in getattr(report, "stability_notes", ()):
        info_sheet.append(["稳定性提示", note])
    info_sheet.append(
        [
            "风险提示",
            "彩票开奖结果具有随机性，模型评分和历史回测不保证未来中奖。",
        ]
    )

    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column
            )
            sheet.column_dimensions[column[0].column_letter].width = min(
                max(12, max_length + 3),
                42,
            )
    workbook.save(path)


def export_backtest_evaluation_xlsx(path: Path, result) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "可信评估摘要"
    summary.append([
        "区域", "回测期数", "模型Brier", "基线Brier", "Brier Skill Score",
        "BSS 95%下限", "BSS 95%上限", "BSS为正概率", "模型平均命中",
        "命中Bootstrap下限", "命中Bootstrap上限", "随机平均命中",
        "随机95%下限", "随机95%上限", "命中提升", "随机实验p值", "结论"
    ])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for label, evaluation in (("前区", result.front_evaluation), ("后区", result.back_evaluation)):
        if evaluation is None:
            continue
        summary.append([
            label, result.evaluated, evaluation.model_brier, evaluation.reference_brier,
            evaluation.brier_skill_score, evaluation.bss_ci_lower, evaluation.bss_ci_upper,
            evaluation.bss_probability_positive, evaluation.model_hit_average,
            evaluation.model_hit_ci_lower, evaluation.model_hit_ci_upper,
            evaluation.random_hit_average, evaluation.random_hit_ci_lower,
            evaluation.random_hit_ci_upper, evaluation.hit_uplift,
            evaluation.random_p_value, evaluation.conclusion,
        ])

    detail = workbook.create_sheet("逐期回测")
    detail.append([
        "期号", "AI前区命中", "AI后区命中", "单次随机前区命中", "单次随机后区命中",
        "模型前区Brier", "基线前区Brier", "模型后区Brier", "基线后区Brier"
    ])
    for cell in detail[1]:
        cell.font = Font(bold=True)
    for row in result.details:
        detail.append([
            row.get("issue"), row.get("model_front_hits"), row.get("model_back_hits"),
            row.get("random_front_hits"), row.get("random_back_hits"),
            row.get("model_front_brier"), row.get("reference_front_brier"),
            row.get("model_back_brier"), row.get("reference_back_brier"),
        ])

    method = workbook.create_sheet("方法说明")
    method.append(["项目", "说明"])
    method.append(["Brier Skill Score", "BSS = 1 - 模型Brier / 均匀概率基线Brier；大于0表示优于基线。"])
    method.append(["Bootstrap置信区间", f"按时间期次有放回重采样 {result.bootstrap_samples:,} 次，报告百分位法95%区间。"])
    method.append(["随机基线重复实验", f"等价超几何随机票实验 {result.random_repeats:,} 次；p值为随机平均命中不低于模型的比例。"])
    method.append(["解释", "置信区间跨越0或随机实验p值较大时，不能认定模型存在稳定优势。"])
    method.append(["风险提示", "彩票开奖具有随机性，历史回测和统计显著性不保证未来中奖。"])
    for cell in method[1]:
        cell.font = Font(bold=True)

    for sheet in workbook.worksheets:
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(12, max_length + 3), 48)
    workbook.save(path)
